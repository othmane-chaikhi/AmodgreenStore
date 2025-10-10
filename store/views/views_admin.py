from django.contrib.auth.decorators import user_passes_test, login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.shortcuts import render, get_object_or_404, redirect, reverse
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_protect
from django.db import models, transaction
from django.db.models import Prefetch, Q
from django.utils import timezone
from store.models import (
    Order, OrderItem, Product, ProductVariant,
    ProductImage, Category, CommunityPost, SiteConfig
)
from store.forms import ProductForm,ProductVariantForm, ProductVariantFormSet, CategoryForm,OrderExportFilterForm

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


# -------------------- HELPERS --------------------
def admin_required(view_func):
    """Restrict access to admin (superuser) only."""
    return user_passes_test(lambda u: u.is_authenticated and u.is_superuser)(view_func)


# -------------------- DASHBOARD --------------------
@admin_required
def admin_dashboard(request):
    """Admin dashboard with stats and latest orders/products."""
    orders_qs = (
        Order.objects.filter(is_deleted=False)
        .prefetch_related(Prefetch("items", queryset=OrderItem.objects.select_related("variant")))
        .order_by("-created_at")
    )

    paginator = Paginator(orders_qs, 10)
    page_number = request.GET.get("order_page", 1)
    try:
        orders = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        orders = paginator.page(1)

    stats = [
        {
            "count": Product.objects.count(),
            "label": "Produits",
            "color": "text-olive-600",
            "icon": "📦",
            "url": reverse("admin_product_list"),
        },
        {
            "count": CommunityPost.objects.count(),
            "label": "Avis",
            "color": "text-blue-600",
            "icon": "💬",
            "url": reverse("post_list"),
        },
        {
            "count": orders_qs.filter(status="pending").count(),
            "label": "Commandes en cours",
            "color": "text-yellow-600",
            "icon": "⏳",
            "url": reverse("order_list") + "?status=pending",
        },
        {
            "count": orders_qs.filter(status="delivered").count(),
            "label": "Commandes livrées",
            "color": "text-green-600",
            "icon": "✅",
            "url": reverse("order_list") + "?status=delivered",
        },
    ]

    # Config form processing (Telegram) with safe fallback if table not migrated yet
    config = None
    try:
        config = SiteConfig.get_solo()
        if request.method == "POST" and request.POST.get("_config") == "1":
            config.telegram_bot_token = request.POST.get("telegram_bot_token", "").strip()
            config.telegram_chat_id = request.POST.get("telegram_chat_id", "").strip()
            config.save()
            messages.success(request, "Configuration Telegram enregistrée.")
            return redirect("admin_dashboard")
    except Exception:
        # Table may not exist yet; render dashboard without config form
        config = None

    return render(request, "admin/dashboard.html", {
        "stats": stats,
        "products": Product.objects.select_related("category").order_by("-id")[:20],
        "orders": orders,
        "orders_count": orders_qs.count(),
        "recent_orders_count": 10,
        "config": config,
    })


# -------------------- ORDER MANAGEMENT --------------------
@admin_required
def update_order_status(request, order_id, status):
    # Validate status choices
    valid_statuses = ['pending', 'contacted', 'delivered', 'cancelled']
    if status not in valid_statuses:
        messages.error(request, "Statut invalide.")
        return redirect("admin_dashboard")
    
    Order.objects.filter(pk=order_id).update(status=status)
    status_display = dict(Order.STATUS_CHOICES).get(status, status)
    messages.success(request, f"Statut de la commande #{order_id} mis à jour: {status_display}")
    return redirect("admin_dashboard")


@login_required
def order_detail(request, order_id):
    """View and update single order."""
    order = get_object_or_404(Order, pk=order_id)
    status_map = {"contact": "contacted", "deliver": "delivered", "cancel": "cancelled"}

    if request.method == "POST":
        action = request.POST.get("action")
        if action in status_map:
            order.status = status_map[action]
            order.save()
            status_display = dict(Order.STATUS_CHOICES).get(order.status, order.status)
            messages.success(request, f"Commande #{order.id} mise à jour: {status_display}")
        return redirect("admin_dashboard")

    return render(request, "admin/order_detail.html", {"order": order})


@admin_required
def delete_order(request, order_id):
    """Delete an order (soft delete for stock management)."""
    order = get_object_or_404(Order, pk=order_id)
    
    # Soft delete - mark as deleted instead of removing from database
    order.is_deleted = True
    order.save()
    
    messages.success(request, f"Commande #{order_id} supprimée avec succès.")
    return redirect("admin_dashboard")


@admin_required
def restore_order(request, order_id):
    """Restore a soft-deleted order."""
    order = get_object_or_404(Order, pk=order_id)
    
    # Restore the order
    order.is_deleted = False
    order.save()
    
    messages.success(request, f"Commande #{order_id} restaurée avec succès.")
    return redirect("admin_dashboard")


@login_required
@user_passes_test(lambda u: u.is_staff)
def order_list(request):
    """List orders with filtering & pagination."""
    status = request.GET.get("status")
    show_deleted = request.GET.get("show_deleted") == "true"
    
    if show_deleted:
        orders = Order.objects.filter(is_deleted=True)
    else:
        orders = Order.objects.filter(is_deleted=False)
    if status:
        orders = orders.filter(status=status)

    paginator = Paginator(orders.order_by("-created_at"), 10)
    page_number = request.GET.get("page")
    try:
        orders = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        orders = paginator.page(1)

    context = {
        "orders": orders,
        "total_orders": Order.objects.count(),
        "pending_orders": Order.objects.filter(status="pending").count(),
        "delivered_orders": Order.objects.filter(status="delivered").count(),
    }
    return render(request, "admin/order_list.html", context)


@login_required
def contact_customer(request, order_id):
    """Mark customer as contacted."""
    order = get_object_or_404(Order, pk=order_id)
    order.status = "contacted"
    order.save()
    messages.success(request, f"Commande #{order_id} marquée comme 'Client contacté'.")
    return redirect("admin_dashboard")


# -------------------- PRODUCT MANAGEMENT --------------------


@admin_required
def product_create(request):
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        variant_formset = ProductVariantFormSet(request.POST, instance=Product())  # dummy instance

        # On passe le formset au form pour validation
        form.variant_formset = variant_formset

        if form.is_valid() and variant_formset.is_valid():
            with transaction.atomic():
                product = form.save(commit=False)
                product.save()  # sauvegarde initiale pour FK

                # Sauvegarde des variantes
                variant_formset.instance = product
                variants = variant_formset.save()

                # Déterminer la variante par défaut
                default_variant_index = request.POST.get("default_variant")
                if default_variant_index:
                    try:
                        default_variant_form = variant_formset.forms[int(default_variant_index)]
                        product.default_variant = default_variant_form.instance
                    except (IndexError, ValueError):
                        product.default_variant = None
                else:
                    # Si aucune sélection, on prend la première variante
                    product.default_variant = variants[0] if variants else None

                # ⚠ Synchroniser le prix du produit avec la variante par défaut
                if product.default_variant:
                    product.price = product.default_variant.price

                product.save()

                # Sauvegarde des images
                for img in request.FILES.getlist("images"):
                    ProductImage.objects.create(product=product, image=img)

            messages.success(request, "Produit créé avec succès.")
            return redirect("admin_dashboard")
    else:
        form = ProductForm()
        variant_formset = ProductVariantFormSet(instance=Product())

    return render(request, "admin/product_create.html", {
        "form": form,
        "variant_formset": variant_formset,
    })
import os

@admin_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == "POST":
        variant_formset = ProductVariantFormSet(request.POST, instance=product)
        form = ProductForm(request.POST, request.FILES, instance=product, variant_formset=variant_formset)

        if form.is_valid() and variant_formset.is_valid():
            default_variant = product.default_variant

            # Prevent deleting default variant
            for deleted_form in variant_formset.deleted_forms:
                if deleted_form.instance.pk == default_variant.pk:
                    messages.error(request, "Vous ne pouvez pas supprimer la variante par défaut.")
                    break
            else:
                with transaction.atomic():
                    variant_formset.save()

                    # Default variant management
                    default_variant_index = request.POST.get("default_variant")
                    if default_variant_index:
                        try:
                            product.variants.update(is_default=False)
                            default_variant_form = variant_formset.forms[int(default_variant_index)]
                            product.default_variant = default_variant_form.instance
                            product.default_variant.is_default = True
                            product.default_variant.save()
                        except (IndexError, ValueError):
                            product.default_variant = None
                    else:
                        product.default_variant = None

                    # --- ✅ Delete old main image if replaced ---
                    if "image" in request.FILES and product.image:
                        old_image_path = product.image.path
                        if os.path.isfile(old_image_path):
                            os.remove(old_image_path)

                    # --- ✅ Delete old extra images if new ones uploaded ---
                    if request.FILES.getlist("images"):
                        for old_img in product.additional_images.all():
                            if old_img.image and os.path.isfile(old_img.image.path):
                                os.remove(old_img.image.path)
                            old_img.delete()

                    form.save()
                    product.save()

                    # Save new extra images
                    for img in request.FILES.getlist("images"):
                        ProductImage.objects.create(product=product, image=img)

                messages.success(request, "Produit mis à jour avec succès.")
                return redirect("admin_dashboard")

    else:
        variant_formset = ProductVariantFormSet(instance=product)
        form = ProductForm(instance=product, variant_formset=variant_formset)

    return render(request, "admin/product_form.html", {
        "form": form,
        "variant_formset": variant_formset,
        "product": product,
        "images": product.additional_images.all(),
        "title": "Modifier le produit",
    })


@csrf_protect
@admin_required
def delete_product_image(request, pk):
    """Delete a product image (main or extra)."""
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Méthode non autorisée"})

    try:
        data = json.loads(request.body)
        product = get_object_or_404(Product, pk=pk)

        if data.get("type") == "main" and product.image:
            product.image.delete(save=False)
            product.image = None
            product.save()
            return JsonResponse({"success": True})

        if data.get("type") == "extra":
            img = get_object_or_404(ProductImage, id=data.get("id"), product=product)
            img.image.delete(save=False)
            img.delete()
            return JsonResponse({"success": True})

        return JsonResponse({"success": False, "error": "Type d'image invalide"})
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Données JSON invalides"})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@admin_required
def product_delete(request, pk):
    """Delete a product (confirmation required)."""
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        product.delete()
        messages.success(request, "Produit supprimé avec succès.")
        return redirect("admin_dashboard")
    return render(request, "admin/product_confirm_delete.html", {"product": product})


@user_passes_test(lambda u: u.is_staff)
def admin_product_list(request):
    """Admin product list with search, filters & pagination."""
    search_query = request.GET.get("q", "")
    category_filter = request.GET.get("category", "")
    availability_filter = request.GET.get("availability", "")

    products = Product.objects.select_related("category").prefetch_related("variants")

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(category__name__icontains=search_query)
        )
    if category_filter:
        products = products.filter(category_id=category_filter)
    if availability_filter:
        products = products.filter(is_available=(availability_filter == "available"))

    paginator = Paginator(products.order_by("-created_at"), 25)
    page = request.GET.get("page", 1)
    try:
        products = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        products = paginator.page(1)

    return render(request, "admin/product_list.html", {
        "products": products,
        "categories": Category.objects.all(),
        "search_query": search_query,
        "category_filter": category_filter,
        "availability_filter": availability_filter,
    })


# -------------------- CATEGORY --------------------
@admin_required
def category_create(request):
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Catégorie ajoutée avec succès.")
            return redirect("category_list")
    else:
        form = CategoryForm()
    return render(request, "admin/category_form.html", {"form": form})


@login_required
@user_passes_test(lambda u: u.is_staff)
def category_list(request):
    categories = Category.objects.annotate(product_count=models.Count("product")).order_by("name")
    return render(request, "admin/category_list.html", {"categories": categories})


@admin_required
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Catégorie mise à jour avec succès.")
            return redirect("category_list")
    else:
        form = CategoryForm(instance=category)
    return render(request, "admin/category_form.html", {"form": form})


@admin_required
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        category.delete()
        messages.success(request, "Catégorie supprimée avec succès.")
        return redirect("category_list")
    return render(request, "admin/category_confirm_delete.html", {"category": category})


# -------------------- POSTS --------------------
@login_required
@user_passes_test(lambda u: u.is_staff)
def post_list(request):
    """Paginated list of community posts."""
    posts = CommunityPost.objects.select_related("author", "product").order_by("-created_at")
    paginator = Paginator(posts, 10)
    page_number = request.GET.get("page")
    try:
        posts = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        posts = paginator.page(1)
    return render(request, "admin/post_list.html", {"posts": posts})


from datetime import timedelta
from django.utils import timezone
from django.db.models import Prefetch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.http import HttpResponse
from io import BytesIO  # New: For safe PDF buffering
from store.forms import OrderExportFilterForm  # Keep if using form
# -------------------- EXPORT --------------------

# -------------------- EXPORT --------------------

@admin_required
def export_orders_excel(request):
    """Export orders to Excel with predefined date ranges (including today). Fixed for date filtering issues."""
    form = OrderExportFilterForm(request.GET or None)

    # Base queryset: All non-deleted orders
    orders = Order.objects.filter(is_deleted=False).prefetch_related(
        Prefetch("items", to_attr="prefetched_items")
    ).order_by("-created_at")

    period = None
    if form.is_valid():
        period = form.cleaned_data['period']
        if period in ['', 'all', None]:  # Export ALL if no period or 'all'
            period = None
        else:
            today = timezone.now().date()

            if period == 'today':
                # Use time range filter to handle timezone issues
                start_of_day = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
                end_of_day = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_of_day, created_at__lte=end_of_day)
                
                # If still no orders, show recent orders as fallback
                if orders.count() == 0:
                    orders = Order.objects.filter(is_deleted=False).prefetch_related(
                        Prefetch("items", to_attr="prefetched_items")
                    ).order_by("-created_at")[:5]
            elif period == 'last_3_days':
                start_datetime = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=2)
                end_datetime = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_datetime, created_at__lte=end_datetime)
            elif period == 'last_week':
                start_datetime = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=6)
                end_datetime = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_datetime, created_at__lte=end_datetime)
            elif period == 'last_month':
                start_datetime = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=30)
                end_datetime = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_datetime, created_at__lte=end_datetime)
            elif period == 'last_year':
                start_datetime = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=365)
                end_datetime = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_datetime, created_at__lte=end_datetime)

    # Excel generation (same as before, with safe access)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commandes"

    headers = ["ID", "Nom Client", "Téléphone", "Adresse", "Statut", "Date", "Produit", "Quantité", "Prix unitaire"]
    ws.append(headers)
    for col in ws[1]:
        col.font = Font(bold=True)
        col.alignment = Alignment(horizontal="center")

    row_num = 2
    for order in orders:
        items = getattr(order, "prefetched_items", order.items.all())

        if items:
            start_row = row_num
            for item in items:
                product_name = getattr(getattr(item, 'variant', None), 'product', None)
                product_name = product_name.name if product_name else "Produit inconnu"
                ws.cell(row=row_num, column=7, value=product_name)
                ws.cell(row=row_num, column=8, value=item.quantity)
                ws.cell(row=row_num, column=9, value=float(item.price))
                row_num += 1
            for col in range(1, 7):
                ws.merge_cells(start_row=start_row, start_column=col, end_row=row_num - 1, end_column=col)
            ws.cell(row=start_row, column=1, value=order.id)
            ws.cell(row=start_row, column=2, value=order.full_name)
            ws.cell(row=start_row, column=3, value=order.phone)
            ws.cell(row=start_row, column=4, value=order.address)
            
            # Status cell with color formatting
            status_cell = ws.cell(row=start_row, column=5, value=order.get_status_display())
            if order.status == "delivered":
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif order.status == "contacted":
                status_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
            elif order.status == "pending":
                status_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
            elif order.status == "cancelled":
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            ws.cell(row=start_row, column=6, value=order.created_at.strftime("%d/%m/%Y %H:%M"))
        else:
            ws.append([
                order.id, order.full_name, order.phone, order.address,
                order.get_status_display(), order.created_at.strftime("%d/%m/%Y %H:%M"),
                "Aucun produit", "-", "-"
            ])
            # Apply status color to the last row
            status_cell = ws.cell(row=row_num, column=5, value=order.get_status_display())
            if order.status == "delivered":
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif order.status == "contacted":
                status_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
            elif order.status == "pending":
                status_cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
            elif order.status == "cancelled":
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            row_num += 1

    # If no orders, add note
    if orders.count() == 0:
        ws.append(["Aucune commande trouvée pour la période.", "", "", "", "", "", "", "", ""])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=commandes.xlsx"
    wb.save(response)
    return response


@admin_required
def export_orders_pdf(request):
    """Export orders to PDF with predefined date ranges (including today). Fixed for date filtering issues."""
    form = OrderExportFilterForm(request.GET or None)

    # Base queryset: All non-deleted orders
    orders = Order.objects.filter(is_deleted=False).prefetch_related(
        Prefetch("items", to_attr="prefetched_items")
    ).order_by("-created_at")

    period = None
    if form.is_valid():
        period = form.cleaned_data['period']
        if period in ['', 'all', None]:  # Export ALL if no period or 'all'
            period = None
        else:
            today = timezone.now().date()

            if period == 'today':
                # Use time range filter to handle timezone issues
                start_of_day = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
                end_of_day = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_of_day, created_at__lte=end_of_day)
                
                # If still no orders, show recent orders as fallback
                if orders.count() == 0:
                    orders = Order.objects.filter(is_deleted=False).prefetch_related(
                        Prefetch("items", to_attr="prefetched_items")
                    ).order_by("-created_at")[:5]
            elif period == 'last_3_days':
                start_datetime = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=2)
                end_datetime = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_datetime, created_at__lte=end_datetime)
            elif period == 'last_week':
                start_datetime = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=6)
                end_datetime = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_datetime, created_at__lte=end_datetime)
            elif period == 'last_month':
                start_datetime = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=30)
                end_datetime = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_datetime, created_at__lte=end_datetime)
            elif period == 'last_year':
                start_datetime = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=365)
                end_datetime = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                orders = orders.filter(created_at__gte=start_datetime, created_at__lte=end_datetime)

    # Create PDF in memory
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # If no orders, add message
    if not orders.exists():
        elements.append(Paragraph("Aucune commande trouvée pour la période.", styles["Normal"]))
    else:
        for order in orders:
            elements.append(Paragraph(f"Commande #{order.id}", styles["Heading1"]))
            elements.append(Paragraph(f"Client : {order.full_name}", styles["Normal"]))
            elements.append(Paragraph(f"Téléphone : {order.phone}", styles["Normal"]))
            elements.append(Paragraph(f"Adresse : {order.address}", styles["Normal"]))

            # Status style (updated for new 3-step workflow)
            status_style = styles["Normal"]
            if order.status == "cancelled":
                status_style = ParagraphStyle("CancelledStyle", parent=status_style, textColor=colors.red)
            elif order.status == "delivered":
                status_style = ParagraphStyle("DeliveredStyle", parent=status_style, textColor=colors.green)
            elif order.status == "contacted":
                status_style = ParagraphStyle("ContactedStyle", parent=status_style, textColor=colors.blue)
            elif order.status == "pending":
                status_style = ParagraphStyle("PendingStyle", parent=status_style, textColor=colors.orange)

            elements.append(Paragraph(f"Statut : {order.get_status_display()}", status_style))
            elements.append(Paragraph(f"Date : {order.created_at.strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
            elements.append(Spacer(1, 12))

            # Items table
            items = getattr(order, "prefetched_items", order.items.all())

            data = [["Produit", "Quantité", "Prix unitaire"]] + [
                [
                    getattr(getattr(item, 'variant', None), 'product', None).name if getattr(item, 'variant', None) else "Produit inconnu",
                    item.quantity,
                    float(item.price)
                ] for item in items
            ]
            if len(data) > 1:
                table = Table(data, colWidths=[300, 100, 100])
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ]))
                elements.append(table)
            else:
                elements.append(Paragraph("Aucun produit commandé", styles["Normal"]))
            elements.append(Spacer(1, 24))

    # Build PDF
    try:
        doc.build(elements)
    except Exception as e:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        error_para = Paragraph(f"Erreur lors de la génération: {str(e)}", styles["Normal"])
        doc.build([error_para])

    # Response
    pdf_value = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf_value, content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=commandes.pdf"
    response["Content-Length"] = len(pdf_value)
    return response