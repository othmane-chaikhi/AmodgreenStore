from django.contrib.auth.decorators import user_passes_test, login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt

from .forms import ProductForm, ConfirmOrderForm, CategoryForm
from .models import Product, Order, CommunityPost, Comment, ProductImage

import json
import openpyxl
from openpyxl.styles import Font, Alignment

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ----------------------------------------
# Décorateur custom pour restreindre aux superusers
# ----------------------------------------
def admin_required(view_func):
    return user_passes_test(lambda u: u.is_authenticated and u.is_superuser)(view_func)


# ----------------------------------------
# DASHBOARD ADMIN
# Affichage des statistiques et liste des produits / commandes paginées
# ----------------------------------------
@admin_required
def admin_dashboard(request):
    total_products = Product.objects.count()
    total_comments = Comment.objects.count()
    total_posts = CommunityPost.objects.count()
    orders_pending = Order.objects.filter(status='pending').count()
    orders_delivered = Order.objects.filter(status='delivered').count()

    products = Product.objects.all().order_by('-id')
    orders_all = Order.objects.filter(is_deleted=False).order_by('-created_at')
    order_page_number = request.GET.get('order_page')
    orders_paginator = Paginator(orders_all, 10)
    orders = orders_paginator.get_page(order_page_number)

    context = {
        'total_products': total_products,
        'total_comments': total_comments,
        'total_posts': total_posts,
        'orders_pending': orders_pending,
        'orders_delivered': orders_delivered,
        'products': products[:20],  # 20 derniers produits
        'orders': orders,
    }
    return render(request, 'admin/dashboard.html', context)


# ----------------------------------------
# Mise à jour du statut d'une commande
# ----------------------------------------
@admin_required
def update_order_status(request, order_id, status):
    order = get_object_or_404(Order, pk=order_id)
    order.status = status
    order.save()
    return redirect('admin_dashboard')


# ----------------------------------------
# Basculer l'approbation d'un commentaire
# ----------------------------------------
@admin_required
def toggle_comment_approval(request, comment_id):
    comment = get_object_or_404(Comment, pk=comment_id)
    comment.is_approved = not comment.is_approved
    comment.save()
    return redirect('admin_dashboard')


# ----------------------------------------
# Création d'un produit avec gestion des images multiples
# ----------------------------------------
@admin_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()

            # Sauvegarde images multiples depuis champ 'image'
            images = request.FILES.getlist('image')
            for img in images:
                ProductImage.objects.create(product=product, image=img)

            messages.success(request, "Produit créé avec succès.")
            return redirect('admin_dashboard')
    else:
        form = ProductForm()

    return render(request, 'admin/product_create.html', {'form': form})


# ----------------------------------------
# Mise à jour produit + suppression images via AJAX
# ----------------------------------------
@admin_required
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)

    # Suppression image via requête JSON (AJAX)
    if request.method == 'POST' and request.content_type == "application/json":
        data = json.loads(request.body)
        img_type = data.get("type")
        img_id = data.get("id")

        if img_type == "extra":
            img_to_delete = get_object_or_404(ProductImage, id=img_id, product=product)
            img_to_delete.delete()
        elif img_type == "main":
            if product.image:
                product.image.delete()
                product.image = None
                product.save()
        return JsonResponse({"success": True})

    # Mise à jour formulaire produit + ajout nouvelles images
    elif request.method == 'POST' and request.content_type.startswith('multipart/form-data'):
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            product = form.save()

            # Ajout images supplémentaires
            images = request.FILES.getlist('images')
            for image in images:
                ProductImage.objects.create(product=product, image=image)

            # Réponse JSON si AJAX
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    "success": True,
                    "new_main_image_url": product.image.url if product.image else None
                })

            messages.success(request, "Produit mis à jour avec succès.")
            return redirect('admin_dashboard')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({"success": False, "errors": form.errors})

    # GET : affichage formulaire avec données produit
    else:
        form = ProductForm(instance=product)

    return render(request, 'admin/product_form.html', {
        'form': form,
        'title': 'Modifier le produit',
        'product': product,
        'images': product.productimage_set.all()
    })


# ----------------------------------------
# Suppression d'une image produit (main ou extra) - URL séparée
# ----------------------------------------
@csrf_exempt
@admin_required
def delete_product_image(request, pk):
    if request.method == "POST":
        data = json.loads(request.body)
        img_id = data.get("id")
        img_type = data.get("type")

        product = get_object_or_404(Product, pk=pk)

        if img_type == "main":
            if product.image:
                product.image.delete()
                product.image = None
                product.save()
                return JsonResponse({"success": True})
            return JsonResponse({"success": False, "error": "Aucune image principale"})

        elif img_type == "extra":
            img = get_object_or_404(ProductImage, id=img_id, product=product)
            img.image.delete()
            img.delete()
            return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Méthode non autorisée"})


# ----------------------------------------
# Suppression complète d'un produit
# ----------------------------------------
@admin_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('admin_dashboard')
    return render(request, 'admin/product_confirm_delete.html', {'product': product})


# ----------------------------------------
# Confirmation d'une commande (statut 'contacted')
# ----------------------------------------
@admin_required
def confirm_order(request, order_id):
    order = get_object_or_404(Order, pk=order_id)

    if request.method == 'POST':
        order.status = 'contacted'
        order.save()
        messages.success(request, f"La commande de {order.full_name} a été confirmée.")
        return redirect('admin_dashboard')

    return render(request, 'admin/order_detail.html', {'order': order})


# ----------------------------------------
# Détail d'une commande + gestion confirmation ou annulation (pour user connecté)
# ----------------------------------------
@login_required
def order_detail(request, order_id):
    order = get_object_or_404(Order, pk=order_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'confirm':
            order.status = 'contacted'  # ou 'confirmed'
            order.save()
            messages.success(request, f"Commande #{order.id} confirmée.")
        elif action == 'cancel':
            order.status = 'cancelled'
            order.save()
            messages.success(request, f"Commande #{order.id} annulée.")
        return redirect('admin_dashboard')

    return render(request, 'admin/order_detail.html', {'order': order})


# ----------------------------------------
# Suppression d'une commande
# ----------------------------------------
@admin_required
def delete_order(request, order_id):
    order = get_object_or_404(Order, pk=order_id)

    if request.method == 'POST':
        order.delete()
        messages.success(request, f"Commande #{order_id} supprimée définitivement.")
        return redirect('admin_dashboard')

    return render(request, 'admin/order_confirm_delete.html', {'order': order})


# ----------------------------------------
# Export commandes en Excel
# ----------------------------------------
@admin_required
def export_orders_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commandes"

    headers = ["ID", "Nom Client", "Téléphone", "Adresse", "Statut", "Date", "Produit", "Quantité"]
    ws.append(headers)

    # Style en-têtes
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for order in Order.objects.all().order_by("-created_at"):
        items = list(order.items.all())
        if items:
            start_row = row_num
            for item in items:
                ws.cell(row=row_num, column=8, value=item.product.name)
                ws.cell(row=row_num, column=9, value=item.quantity)
                row_num += 1

            # Fusion des colonnes 1 à 7 sur les lignes dupliquées
            for col in range(1, 8):
                ws.merge_cells(start_row=start_row, start_column=col, end_row=row_num - 1, end_column=col)

            # Remplissage des infos de la commande
            ws.cell(row=start_row, column=1, value=order.id)
            ws.cell(row=start_row, column=2, value=order.full_name)
            ws.cell(row=start_row, column=3, value=order.phone)
            ws.cell(row=start_row, column=5, value=order.address)
            ws.cell(row=start_row, column=6, value=order.get_status_display())
            ws.cell(row=start_row, column=7, value=order.created_at.strftime("%d/%m/%Y %H:%M"))

        else:
            # Commande sans produit
            ws.cell(row=row_num, column=1, value=order.id)
            ws.cell(row=row_num, column=2, value=order.full_name)
            ws.cell(row=row_num, column=3, value=order.phone)
            ws.cell(row=row_num, column=5, value=order.address)
            ws.cell(row=row_num, column=6, value=order.get_status_display())
            ws.cell(row=row_num, column=7, value=order.created_at.strftime("%d/%m/%Y %H:%M"))
            ws.cell(row=row_num, column=8, value="Aucun produit")
            ws.cell(row=row_num, column=9, value="-")
            row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=commandes.xlsx'
    wb.save(response)
    return response


# ----------------------------------------
# Export commandes en PDF
# ----------------------------------------
@admin_required
def export_orders_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=commandes.pdf'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    style_title = styles["Heading1"]
    style_normal = styles["Normal"]

    for order in Order.objects.all().order_by("-created_at"):
        elements.append(Paragraph(f"Commande #{order.id}", style_title))
        elements.append(Paragraph(f"Client : {order.full_name}", style_normal))
        elements.append(Paragraph(f"Téléphone : {order.phone}", style_normal))
        elements.append(Paragraph(f"Adresse : {order.address}", style_normal))

        # Style statut selon état
        status_text = f"Statut : {order.get_status_display()}"
        if order.status == 'cancelled':
            status_style = ParagraphStyle(
                "CancelledStyle",
                parent=style_normal,
                textColor=colors.red
            )
            elements.append(Paragraph(status_text, status_style))
        elif order.status == 'delivered':
            status_style = ParagraphStyle(
                "DeliveredStyle",
                parent=style_normal,
                textColor=colors.green
            )
            elements.append(Paragraph(status_text, status_style))
        else:
            elements.append(Paragraph(status_text, style_normal))

        elements.append(Paragraph(f"Date : {order.created_at.strftime('%d/%m/%Y %H:%M')}", style_normal))
        elements.append(Spacer(1, 12))

        # Tableau produits
        data = [["Produit", "Quantité"]]
        for item in order.items.all():
            data.append([item.product.name, item.quantity])

        if len(data) > 1:
            table = Table(data, colWidths=[300, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("Aucun produit commandé", style_normal))

        elements.append(Spacer(1, 24))  # Espace entre commandes

    doc.build(elements)
    return response


# ----------------------------------------
# Création de catégorie produit
# ----------------------------------------
@admin_required
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Catégorie ajoutée avec succès.")
            return redirect('admin_dashboard')
    else:
        form = CategoryForm()
    return render(request, 'admin/category_form.html', {'form': form})
